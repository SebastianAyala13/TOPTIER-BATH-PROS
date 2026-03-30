#!/usr/bin/env python3
"""Inspect Excel file structure for debugging."""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed")
    sys.exit(1)

if len(sys.argv) < 2:
    print("Usage: python inspect_excel.py <excel_file>")
    sys.exit(1)

excel_path = sys.argv[1]
excel_file = Path(excel_path)

if not excel_file.exists():
    print(f"File not found: {excel_path}")
    sys.exit(1)

try:
    wb = load_workbook(excel_file)
    print(f"✓ Loaded workbook: {excel_path}")
    print(f"\n📄 Available sheets: {wb.sheetnames}\n")
    
    # Check if Zips-Bath exists
    if "Zips-Bath" in wb.sheetnames:
        ws = wb["Zips-Bath"]
        print(f"📋 Sheet: Zips-Bath")
        print(f"  Dimensions: {ws.dimensions}")
        
        # Check columns with data
        print(f"\n  First 10 rows (all columns):")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
            values = [cell.value for cell in row if cell.value is not None]
            if values:
                print(f"    Row {row_idx}: {values}")
        
        # Check column F specifically
        print(f"\n  Column F content (first 20 rows):")
        col_f_values = []
        for row_idx, cell in enumerate(ws.iter_cols(min_col=6, max_col=6, values_only=False), start=1):
            if row_idx > 20:
                break
            value = cell[0].value
            if value is not None:
                col_f_values.append(f"    Row {row_idx}: {value}")
        
        if col_f_values:
            for v in col_f_values:
                print(v)
        else:
            print("    (Column F appears to be empty)")
    else:
        print("❌ Sheet 'Zips-Bath' not found in workbook")
        
except Exception as e:
    print(f"❌ Error: {e}")
    sys.exit(1)
