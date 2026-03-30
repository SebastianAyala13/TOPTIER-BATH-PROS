#!/usr/bin/env python3
"""
Generate authorized zip codes TypeScript file from Excel.

This script reads zip codes from a specific Excel sheet and column,
and generates src/lib/authorizedZipCodes.ts with AUTHORIZED_ZIP_CODES Set
and isValidZipCode validation function.

Usage:
    python scripts/generate_authorized_zip_codes.py <path_to_excel>

Example:
    python scripts/generate_authorized_zip_codes.py data/zipcodes.xlsx
"""

import sys
import re
from pathlib import Path
from typing import Set

# Try importing openpyxl; provide helpful error if missing
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not found. Install it with:")
    print("  pip install openpyxl")
    sys.exit(1)


def extract_zip_codes(excel_path: str, sheet_name: str = "Zips-Bath", column_header: str = "Zips exl landing") -> Set[str]:
    """
    Extract and clean zip codes from Excel file.
    
    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the worksheet (default: "Zips-Bath")
        column_header: Column header to search for (default: "Zips exl landing")
    
    Returns:
        Set of cleaned and deduplicated zip codes
    
    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If sheet or column not found, or result is empty
    """
    excel_file = Path(excel_path)
    
    # Validate file exists
    if not excel_file.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    try:
        workbook = load_workbook(excel_file, data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to load Excel file: {e}")
    
    # Validate sheet exists
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(
            f"Sheet '{sheet_name}' not found in workbook.\n"
            f"Available sheets: {available}"
        )
    
    worksheet = workbook[sheet_name]
    
    # Find column index by header in first row
    col_index = None
    col_letter = None
    
    for cell in worksheet[1]:
        if cell.value and str(cell.value).strip() == column_header:
            col_index = cell.column
            col_letter = cell.column_letter
            break
    
    if col_index is None:
        # Get available headers for error message
        available_headers = [cell.value for cell in worksheet[1] if cell.value is not None]
        raise ValueError(
            f"Column header '{column_header}' not found in {sheet_name}.\n"
            f"Available headers: {available_headers}"
        )
    
    # Extract zip codes from column
    zip_codes: Set[str] = set()
    skipped = 0
    duplicates = 0
    max_row = worksheet.max_row
    
    for row_idx in range(2, max_row + 1):
        cell = worksheet.cell(row=row_idx, column=col_index)
        value = cell.value
        
        if value is None or value == "":
            skipped += 1
            continue
        
        # Convert to string and clean
        zip_str = str(value).strip()
        
        # Validate: must be 5 digits
        if not re.match(r'^\d{5}$', zip_str):
            # Try to extract 5 digits if it's part of a longer string
            match = re.search(r'(\d{5})', zip_str)
            if match:
                zip_str = match.group(1)
            else:
                skipped += 1
                continue
        
        # Check for duplicates
        if zip_str in zip_codes:
            duplicates += 1
        else:
            zip_codes.add(zip_str)
    
    if not zip_codes:
        raise ValueError(
            f"No valid zip codes found in {sheet_name}!{col_letter}.\n"
            f"Ensure the column exists and contains 5-digit zip codes."
        )
    
    # Print summary
    print(f"\n✓ Extracted zip codes from {sheet_name}!{col_letter}")
    print(f"  Column header: '{column_header}'")
    print(f"  Total rows scanned: {max_row - 1}")
    print(f"  Valid & unique: {len(zip_codes)}")
    print(f"  Skipped (empty/invalid): {skipped}")
    print(f"  Duplicates removed: {duplicates}")
    
    return zip_codes


def generate_typescript_file(zip_codes: Set[str], output_path: str) -> None:
    """
    Generate TypeScript file with AUTHORIZED_ZIP_CODES and isValidZipCode.
    
    Args:
        zip_codes: Set of zip codes to write
        output_path: Path to write the TypeScript file
    """
    # Sort for consistency and readability
    sorted_zips = sorted(zip_codes)
    
    # Generate zip codes array
    zips_array = ",\n  ".join(f"'{z}'" for z in sorted_zips)
    
    # Generate the TypeScript content
    typescript_content = f"""export const AUTHORIZED_ZIP_CODES = new Set([
  {zips_array}
]);

export const isValidZipCode = (zipCode: string): boolean => {{
  const cleanZip = zipCode.trim();
  return AUTHORIZED_ZIP_CODES.has(cleanZip);
}};
"""
    
    # Write to file
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    output_file.write_text(typescript_content)
    
    print(f"\n✓ Generated {output_path}")
    print(f"  Total zip codes in Set: {len(sorted_zips)}")


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python scripts/generate_authorized_zip_codes.py <path_to_excel>")
        print("       [--sheet SHEET_NAME] [--header COLUMN_HEADER]")
        print("\nExample:")
        print("  python scripts/generate_authorized_zip_codes.py data/zipcodes.xlsx")
        print("  python scripts/generate_authorized_zip_codes.py data/zipcodes.xlsx --sheet 'Zips-Bath' --header 'Zips exl landing'")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    sheet_name = "Zips-Bath"
    column_header = "Zips exl landing"
    
    # Parse optional arguments
    for i, arg in enumerate(sys.argv[2:], start=2):
        if arg == "--sheet" and i + 1 < len(sys.argv):
            sheet_name = sys.argv[i + 1]
        elif arg == "--header" and i + 1 < len(sys.argv):
            column_header = sys.argv[i + 1]
    
    output_path = "src/lib/authorizedZipCodes.ts"
    
    try:
        print(f"\n📋 Reading Excel: {excel_path}")
        zip_codes = extract_zip_codes(excel_path, sheet_name, column_header)
        
        print(f"\n💾 Writing TypeScript file: {output_path}")
        generate_typescript_file(zip_codes, output_path)
        
        print(f"\n✅ Success! Run 'npm run build' to validate.\n")
        
    except (FileNotFoundError, ValueError) as e:
        print(f"\n❌ Error: {e}\n")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
