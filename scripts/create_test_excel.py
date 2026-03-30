#!/usr/bin/env python3
"""Create a test Excel file with sample zip codes for testing."""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Zips-Bath"

# Add header in column F
ws['F1'] = "Zips exl landing"

# Sample zip codes for testing
sample_zips = [
    '01011', '01012', '01026', '01029', '01032', '01050', '01070', '01084',
    '01096', '01098', '01201', '01202', '01203', '01220', '01222', '01223',
    '01224', '01225', '01226', '01227', '01229', '01230', '01235', '01236',
    '01237', '01238', '01240', '01242', '01243', '01244', '01245', '01247',
    '01252', '01253', '01254', '01255', '01256', '01257', '01258', '01259',
    '01260', '01262', '01263', '01264', '01266', '01267', '01270', '01330',
    '02101', '02102', '02103', '02104', '02105', '02106', '02107', '02108',
    '02109', '02110', '02111', '02112', '02113', '02114', '02115', '02116'
]

# Add sample zips to column F (starting from row 2)
for idx, zip_code in enumerate(sample_zips, start=2):
    ws[f'F{idx}'] = zip_code

# Save the test file
output_path = Path("data/test_zipcodes.xlsx")
output_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(output_path)

print(f"✓ Created test Excel file: {output_path}")
print(f"  Sheet: {ws.title}")
print(f"  Column: F (Zips exl landing)")
print(f"  Sample zip codes: {len(sample_zips)}")
